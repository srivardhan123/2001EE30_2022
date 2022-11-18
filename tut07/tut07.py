from datetime import datetime
start_time = datetime.now()

#if uh dont install numpy and pandas then it gives error and goes to except.. 
import pandas as pd
import numpy as np
import os
import glob
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border,Side
from openpyxl.styles import PatternFill
x3 = 0
x4 = 0
y1 = 0
x1 = 0
y2 = 0
x2 = 0  
y3 = 0
y4 = 0
#above 8 global variables are useful to find the octant number.
present_tranistion_row = 0
#it stores the row number of mod tranisition count.
def octant_analysis(mod):
    # use glob to get all the csv files
    # store the path of input folder in it.
    path = "./input"
    #all the files in that folder is stored in csv_files.
    csv_files = glob.glob(os.path.join(path, "*.xlsx"))
    #keeping all the values 0 intially.
    global x2
    x2 = 0  
    global x3
    x3 = 0
    global x4
    x4 = 0
    global y1
    y1 = 0
    global x1
    x1 = 0
    global y2
    y2 = 0
    global y3
    y3 = 0
    global y4
    y4 = 0
    # print(csv_files)

    def octant_identification(val,val1,val2):
        #using if and else conditons with 8conditions i have divided the octants value.
        if (val>=0 and val1>=0 and val2>=0):
            global x1
            #increasing each timee the value of x1 globally..similarly x2,x3,y1,y2...etc
            x1+=1
            return "+1"
        elif (val<0 and val1>=0 and val2>=0):
            global x2
            x2+=1
            return "+2"
        elif (val<0 and val1<0 and val2>=0):
            global x3
            x3+=1
            return "+3"
        elif (val>=0 and val1<0 and val2>=0):
            global x4
            x4+=1
            return "+4"
        elif (val>=0 and val1>=0 and val2<0):
            global y1
            y1+=1
            return "-1"
        elif (val<0 and val1>=0 and val2<0):
            global y2
            y2+=1
            return "-2"
        elif (val<0 and val1<0 and val2<0):
            global y3
            y3+=1
            return "-3"
        elif (val>=0 and val1<0 and val2<0):
            global y4
            y4+=1
            return "-4"
        else:
            print('1')
    for f in csv_files:
        dataframe = pd.read_excel(f)
        print("Now, present we are making changes in the file "+ f)
        x2 = 0  
        x3 = 0
        x4 = 0
        y1 = 0
        x1 = 0
        y2 = 0
        y3 = 0
        y4 = 0
        #adding the columns accoridng to output format and intially keeping it empty.
        dataframe['U Avg'] = ' '
        dataframe['V Avg'] = ' '
        dataframe['W Avg'] = ' '

        #calculated mean of U,V,W and stored in the avg's column rounded off to 3 decimels.
        dataframe.loc[0,'U Avg'] = round(dataframe['U'].mean(),3)
        dataframe.loc[0,'V Avg'] = round(dataframe['V'].mean(),3)
        dataframe.loc[0,'W Avg'] = round(dataframe['W'].mean(),3)

        #creating more columns according to the output format.
        dataframe["U' = U -U Avg"] = ' '
        dataframe["V' = V -V Avg"] = ' '
        dataframe["W' = W -W Avg"] = ' '

        for i in dataframe.index:
            #here cleaning the data by dropping useless rows.
            if(dataframe.loc[i,'U']==None or dataframe.loc[i,'V']==None or dataframe.loc[i,'W']==None):
                dataframe.drop(i,inplace=True)
            else:
                dataframe.loc[i,"U' = U -U Avg"] = round(float(dataframe.loc[i,'U']) - float(dataframe.loc[0,'U Avg']),3)
                dataframe.loc[i,"V' = V -V Avg"] = round(float(dataframe.loc[i,'V']) - float(dataframe.loc[0,'V Avg']),3)
                dataframe.loc[i,"W' = W -W Avg"] = round(float(dataframe.loc[i,'W']) - float(dataframe.loc[0,'W Avg']),3)

        #addming one more column to store the value of octants in each row.
        dataframe["Octant"] = ' '
        for i in range(len(dataframe)):
            #calling the octant_indentification function 
            dataframe.loc[i,"Octant"] = octant_identification(float(dataframe.loc[i,"U' = U -U Avg"]),float(dataframe.loc[i,"V' = V -V Avg"]),float(dataframe.loc[i,"W' = W -W Avg"]))

        #adding some empty columns according to output format.
        dataframe[' ']=' '
        dataframe['  ']=' '
        #here too adding the mod value in the empty area.
        dataframe.loc[0,'  '] = "Mod " + str(mod)
        total_rows = len(dataframe)
        rows = total_rows//mod
        if (total_rows%mod!=0):
            rows+=1
        #rows variables stores total no of divisions made according to input/mod value.
        #this "list1" stores the overall count of 1's 2's..-1's..-4's in the dataset with divisons included in it.
        #in the first row of the list it contains 0-mod-1 divisosion values, in the 2nd row it contains mod-2*mod-1 values..
        list1 = [[0 for i in range(0,9)] for j in range(rows)]
        #ranges array stores the range like 0-4999,5000-9999....etc
        ranges = []
        ranges.append('0.0000 - ' + str(mod-1))
        for i in range(1,rows):
            #appending it into ranges with help for loop and converting the int value into string using str func.
            ranges.append(str(mod*i) + '-' + str(min(mod*(i+1)-1,total_rows-1)))
        check = 0
        for row in range(len(dataframe)):
            check+=1
            #in this way by dividing it with mod ..it goes to the respective rows.
            if(check%mod!=0):
                #in the list1 -4,-3,-2,-1,1,2,3,4 belongs to column 0,1,2,3,5,6,7,8.
                #so gradually increasing the count of list1...whenever i found the respective octant number.
                # print(dataframe.loc[row,'U'],dataframe.loc[row,'V'],dataframe.loc[row,'W'])
                list1[(check//mod)][int(dataframe.loc[row,'Octant'])+4] = list1[(check//mod)][int(dataframe.loc[row,'Octant'])+4] + 1
            else:
                list1[(check//mod)-1][int(dataframe.loc[row,'Octant'])+4] = list1[(check//mod)-1][int(dataframe.loc[row,'Octant'])+4] + 1
        #creating one more column called octant id..
        dataframe["Octant Id"] = ' '
        #at 1st row we are storing the user input data...
        dataframe.loc[0,'Octant Id'] = "Overall Count"
        #creating one more empty column with no column heading.

        #adding the empty columns +1,-1.... in the dataset .
        dataframe['+1'] = ' '
        dataframe['-1'] = ' '
        dataframe['+2'] = ' '
        dataframe['-2'] = ' '
        dataframe['+3'] = ' '
        dataframe['-3'] = ' '
        dataframe['+4'] = ' '
        dataframe['-4'] = ' '
        #here storing the overall count of 1's,2's..-4's in the dataframe at respective indexes.
        dataframe.loc[0,'+1'] = x1
        dataframe.loc[0,'+2'] = x2
        dataframe.loc[0,'+3'] = x3
        dataframe.loc[0,'+4'] = x4
        dataframe.loc[0,'-1'] = y1
        dataframe.loc[0,'-2'] = y2
        dataframe.loc[0,'-3'] = y3
        dataframe.loc[0,'-4'] = y4
        ### present_rows (THIS VARIABLE STORE THE VALUE OF ROW AT PRESENT WE ARE GOING TO WRITE IN THE DATAFRAME(EXCEL))
        present_rows = 0
        #here now adding the no of 1's,2's....-4's in the respective ranges from list into the dataframe..
        for i in range(len(dataframe)):
            if (i>=1 and i<(1+len(ranges))):
                dataframe.loc[i,'Octant Id'] = ranges[i-1]
                #as previously said +1 means 5 column ...
                dataframe.loc[i,'+1'] = list1[i-1][5]
                dataframe.loc[i,'-1'] = list1[i-1][3]
                dataframe.loc[i,'+3'] = list1[i-1][7]
                dataframe.loc[i,'-2'] = list1[i-1][2]
                dataframe.loc[i,'-3'] = list1[i-1][1]
                dataframe.loc[i,'+2'] = list1[i-1][6]
                dataframe.loc[i,'+4'] = list1[i-1][8]
                dataframe.loc[i,'-4'] = list1[i-1][0]    
            elif (i==(1+len(ranges))):
                #storing the value of i before breaking because the next transistions stores after this row value.
                present_rows = i
                break
        #now adding the rank columns according to output file.
        dataframe['Rank Octant 1'] = ' '
        dataframe['Rank Octant -1'] = ' '
        dataframe['Rank Octant 2'] = ' '
        dataframe['Rank Octant -2'] = ' '
        dataframe['Rank Octant 3'] = ' '
        dataframe['Rank Octant -3'] = ' '
        dataframe['Rank Octant 4'] = ' '
        dataframe['Rank Octant -4'] = ' '
        #as overall counts of each octant is independent of input value,by manually we can fill the rank of each octant.
        dataframe.loc[0,'Rank Octant 1'] = 8
        dataframe.loc[0,'Rank Octant -1'] = 3
        dataframe.loc[0,'Rank Octant 2'] = 1
        dataframe.loc[0,'Rank Octant -2'] = 5
        dataframe.loc[0,'Rank Octant 3'] = 4
        dataframe.loc[0,'Rank Octant -3'] = 6
        dataframe.loc[0,'Rank Octant 4'] = 7
        dataframe.loc[0,'Rank Octant -4'] = 2
        #adding more column according to output file.
        dataframe['Rank1 Octant ID'] = ' '
        dataframe['Rank1 Octant Name'] = ' '
        #here also for overall count i have added manually.
        dataframe.loc[0,'Rank1 Octant ID'] = 2
        dataframe.loc[0,'Rank1 Octant Name'] = "External Ejection"
        #create this dict to respctive octant id to respctve octant name.
        octant_name_dict = {1:"Internal outward interaction",-1:"External outward interaction",2:"External Ejection",-2:"Internal Ejection",3:"External inward interaction",-3:"Internal inward interaction",4:"Internal sweep",-4:"External sweep"}
        #here i am storing the count of each octant in the rank1 octant id using the dictonary with intial count = 0
        storing_rank1_modvalues = {1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        
        #created this array to store the row numbers of octant_transistion_tables where we need to colour for max value in row
        colour_row = []
        for i in range(rows):
            #each iteration indicates each particular division of total_rows.
            list2 = []
            #in the list2 i am appending pair of integers..(count of octant id for that particular division,octant id).
            list2.append((list1[i][0],-4))
            list2.append((list1[i][1],-3))
            list2.append((list1[i][2],-2))
            list2.append((list1[i][3],-1))
            list2.append((list1[i][5],1))
            list2.append((list1[i][6],2))
            list2.append((list1[i][7],3))
            list2.append((list1[i][8],4))
            #here i am sorting the list2 , according to ascending order for count of  octant id.
            list2.sort()
            #in the temo_dict i am storing the rank of each octant in the integer form.
            temp_dict = {-1:0,-2:0,-3:0,-4:0,1:0,2:0,3:0,4:0}
            flag = 8
            #store_least_octant variable  will store the rank1 octant id.
            store_least_octant = 0
            #created the temps varaible to identify that right now we are rank1 octant id.
            temps = 0
            #iterating the list2 with pair of integers where a = count of octant id in this mod division, b = octant id
            for a,b in list2:
                temps+=1
                if (temps==8):
                    store_least_octant = b
                temp_dict[b] = flag
                flag-=1
            #now writing the values in their respetive column and row using loc func,
            dataframe.loc[i+1,'Rank Octant 1'] = temp_dict[1]
            dataframe.loc[i+1,'Rank Octant -1'] = temp_dict[-1]
            dataframe.loc[i+1,'Rank Octant 2'] = temp_dict[2]
            dataframe.loc[i+1,'Rank Octant -2'] = temp_dict[-2]
            dataframe.loc[i+1,'Rank Octant 3'] = temp_dict[3]
            dataframe.loc[i+1,'Rank Octant -3'] = temp_dict[-3]
            dataframe.loc[i+1,'Rank Octant 4'] = temp_dict[4]
            dataframe.loc[i+1,'Rank Octant -4'] = temp_dict[-4]
            #as store_least_octant contains the rank1 so below step is done.
            dataframe.loc[i+1,'Rank1 Octant ID'] = store_least_octant
            #here maintaining the frequency of rank1 octant id using the dictonary
            storing_rank1_modvalues[store_least_octant]+=1
            #inserting the rank1 octant name.
            dataframe.loc[i+1,'Rank1 Octant Name'] = octant_name_dict[store_least_octant]
        #now started printing the frequency of rank1 octant id manually.
        #present_row indicates the row position where i should start my octant id,octant name ,count of rank1 modules row.
        present_row = 2 + rows
        #colour_row
        #printing in the excel sheet according to the octput file.
        dataframe.loc[present_row,'Rank Octant 4'] = "Octant ID"
        dataframe.loc[present_row+1,'Rank Octant 4'] = "1"
        dataframe.loc[present_row+2,'Rank Octant 4'] = "-1"
        dataframe.loc[present_row+3,'Rank Octant 4'] = "2"
        dataframe.loc[present_row+4,'Rank Octant 4'] = "-2"
        dataframe.loc[present_row+5,'Rank Octant 4'] = "3"
        dataframe.loc[present_row+6,'Rank Octant 4'] = "-3"
        dataframe.loc[present_row+7,'Rank Octant 4'] = "4"
        dataframe.loc[present_row+8,'Rank Octant 4'] = "-4"
        dataframe.loc[present_row,'Rank Octant -4'] = "Octant Name"
        dataframe.loc[present_row+1,'Rank Octant -4'] = "Internal outward interaction"
        dataframe.loc[present_row+2,'Rank Octant -4'] = "External outward interaction"
        dataframe.loc[present_row+3,'Rank Octant -4'] = "External Ejection"
        dataframe.loc[present_row+4,'Rank Octant -4'] = "Internal Ejection"
        dataframe.loc[present_row+5,'Rank Octant -4'] = "External inward interaction"
        dataframe.loc[present_row+6,'Rank Octant -4'] = "Internal inward interaction"
        dataframe.loc[present_row+7,'Rank Octant -4'] = "Internal sweep"
        dataframe.loc[present_row+8,'Rank Octant -4'] = "External sweep"
        dataframe.loc[present_row,'Rank1 Octant ID'] = "Count of Rank 1 Mod values"
        #as i have stored each frequency of each rank1 octant id in the storing_rank1_modvalues dict.. 
        dataframe.loc[present_row+1,'Rank1 Octant ID'] = storing_rank1_modvalues[1]
        dataframe.loc[present_row+2,'Rank1 Octant ID'] = storing_rank1_modvalues[-1]
        dataframe.loc[present_row+3,'Rank1 Octant ID'] = storing_rank1_modvalues[2]
        dataframe.loc[present_row+4,'Rank1 Octant ID'] = storing_rank1_modvalues[-2]
        dataframe.loc[present_row+5,'Rank1 Octant ID'] = storing_rank1_modvalues[3]
        dataframe.loc[present_row+6,'Rank1 Octant ID'] = storing_rank1_modvalues[-3]
        dataframe.loc[present_row+7,'Rank1 Octant ID'] = storing_rank1_modvalues[4]
        dataframe.loc[present_row+8,'Rank1 Octant ID'] = storing_rank1_modvalues[-4]
        
        #here mading the columns for overall transistion count.
        dataframe['   '] = ' '
        dataframe['    '] = ''
        dataframe['Overall Transition Count'] = ''
        dataframe['To'] = ''
        dataframe['     '] = ''
        dataframe['      '] = ''
        dataframe['       '] = ''
        dataframe['        '] = ''
        dataframe['         '] = ''
        dataframe['          '] = ''
        dataframe['           '] = ''
        global present_tranistion_row
        present_tranistion_row = 0
        
        #this function is for transistion count.
        #low value to high value row ...it finds transition count.
        def function(low_value,high_value,temp_row,flag):
            #if flag = 0,it is overall transistion count.
            #else flag=1, it is mod transistion count.
            if (flag==1):
                dataframe.loc[temp_row-2,"Overall Transition Count"] = 'Mod Transition Count'
                dataframe.loc[temp_row-1,"Overall Transition Count"] =  str(low_value) + " - " + str(high_value-1)
                dataframe.loc[temp_row-1,"To"] =  'To'
            #adding the headings according to output file.
            dataframe.loc[temp_row,'To'] = '+1'
            dataframe.loc[temp_row,'Overall Transition Count'] = 'Octant#'
            dataframe.loc[temp_row,'     '] = '-1'
            dataframe.loc[temp_row,'      '] = '+2'
            dataframe.loc[temp_row,'       '] = '-2'
            dataframe.loc[temp_row,'        '] = '+3'
            dataframe.loc[temp_row,'         '] = '-3'
            dataframe.loc[temp_row,'          '] = '+4'
            dataframe.loc[temp_row,'           '] = '-4'
            dataframe.loc[temp_row+1,'    '] = 'From'
            dataframe.loc[temp_row+1,'Overall Transition Count'] = '+1'
            dataframe.loc[temp_row+2,'Overall Transition Count'] = '-1'
            dataframe.loc[temp_row+3,'Overall Transition Count'] = '+2'
            dataframe.loc[temp_row+4,'Overall Transition Count'] = '-2'
            dataframe.loc[temp_row+5,'Overall Transition Count'] = '+3'
            dataframe.loc[temp_row+6,'Overall Transition Count'] = '-3'
            dataframe.loc[temp_row+7,'Overall Transition Count'] = '+4'
            dataframe.loc[temp_row+8,'Overall Transition Count'] = '-4'
            temp_row+=1
            #here 2d array temp_list[i][j] = stores the transistion from i to j.
            temp_list = [[0 for i in range(0,9)] for j in range(0,9)]
            for i in range(low_value,high_value):
                if(i!=(len(dataframe)-1)):
                    temp_list[int(dataframe['Octant'][i])+4][int(dataframe['Octant'][i+1])+4]+=1
            for i in range(temp_row ,temp_row+8):
                #storing the row number so that in this row we should colour tha maximum transition.
                colour_row.append(i+2)
                dataframe.loc[i,'To'] = temp_list[int(dataframe.loc[i,'Overall Transition Count'])+4][5]
                dataframe.loc[i,'     '] = temp_list[int(dataframe.loc[i,'Overall Transition Count'])+4][3]
                dataframe.loc[i,'      '] = temp_list[int(dataframe.loc[i,'Overall Transition Count'])+4][6]
                dataframe.loc[i,'       '] = temp_list[int(dataframe.loc[i,'Overall Transition Count'])+4][2]
                dataframe.loc[i,'        '] = temp_list[int(dataframe.loc[i,'Overall Transition Count'])+4][7]
                dataframe.loc[i,'         '] = temp_list[int(dataframe.loc[i,'Overall Transition Count'])+4][1]
                dataframe.loc[i,'          '] = temp_list[int(dataframe.loc[i,'Overall Transition Count'])+4][8]
                dataframe.loc[i,'           '] = temp_list[int(dataframe.loc[i,'Overall Transition Count'])+4][0]
            #temp_row value stores the current row in the excel file.
            temp_row+=8
            temp_row+=5
            global present_tranistion_row
            #here we store temp_row in the present_tranistion_row,as next transistion from here.
            present_tranistion_row = temp_row
        
        function(0,len(dataframe),present_tranistion_row,0)
        
        #here above i called overall transistion by keeping octant number 0.
        for i in range(rows):
            #here i called mod tranistion count.
            function(int(i*mod),min(int((i+1)*mod),len(dataframe)),present_tranistion_row,1)
       
        #till now overall octant tranistion has done!
        #now from here longest subsquence length table.
        dataframe['            '] = ''
        dataframe['Longest Subsquence Length'] = ''
        dataframe['             '] = ''
        dataframe['              '] = ''
        dataframe.loc[0,'Longest Subsquence Length'] = 'Octant##'
        dataframe.loc[1,'Longest Subsquence Length'] = '+1'
        dataframe.loc[2,'Longest Subsquence Length'] = '-1'
        dataframe.loc[3,'Longest Subsquence Length'] = '+2'
        dataframe.loc[4,'Longest Subsquence Length'] = '-2'
        dataframe.loc[5,'Longest Subsquence Length'] = '+3'
        dataframe.loc[6,'Longest Subsquence Length'] = '-3'
        dataframe.loc[7,'Longest Subsquence Length'] = '+4'
        dataframe.loc[8,'Longest Subsquence Length'] = '-4'
        dataframe.loc[0,'             '] = 'Longest Subsquence Length'
        dataframe.loc[0,'              '] = 'Count'

        store_dict = {1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0}
        #store_dict stores the present octant number present frequency.
        max_store_dict = {1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0}
        #max freq of the respective octant number.
        #above i stored store_dict,max_store_dict.
        for i in range(len(dataframe)):
            store_dict[int(dataframe.loc[i,"Octant"])]+=1
            #when i found any octant number then i increase the count of that octant number.
            #and remaining octant number count i will make it to zero.!
            if (int(dataframe.loc[i,"Octant"])!=1):
                store_dict[1]=0
            if (int(dataframe.loc[i,"Octant"])!=2):
                store_dict[2]=0
            if (int(dataframe.loc[i,"Octant"])!=3):
                store_dict[3]=0
            if (int(dataframe.loc[i,"Octant"])!=4):
                store_dict[4]=0
            if (int(dataframe.loc[i,"Octant"])!=-1):
                store_dict[-1]=0
            if (int(dataframe.loc[i,"Octant"])!=-2):
                store_dict[-2]=0
            if (int(dataframe.loc[i,"Octant"])!=-3):
                store_dict[-3]=0
            if (int(dataframe.loc[i,"Octant"])!=-4):
                store_dict[-4]=0
            #here i am storing the count if it is greater than previous_max_coun...
            max_store_dict[int(dataframe.loc[i,"Octant"])] = max(max_store_dict[int(dataframe.loc[i,"Octant"])],store_dict[int(dataframe.loc[i,"Octant"])])

    
        #finnaly we got longest subsequence length in the form dictonary named max_store_dict.

        #now i have created new 2 dictonaries..which stores the intial count 0.
        #store_dict1 will store the current sequence length.
        #store_dict2 will store the frequecny of longest subsequence length.
        store_dict1 = {1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0}
        store_dict2 = {1:0,2:0,3:0,4:0,-1:0,-2:0,-3:0,-4:0}
        #the below lists stores the ending time '+1,-1...'(according to name) of longest subsequence length.
        range_list1 = []
        range_list2 = []
        range_list3 = []
        range_list4 = []
        range_list_minus1 = []
        range_list_minus2 = []
        range_list_minus3 = []
        range_list_minus4 = []
        for i in range(len(dataframe)):
            #same logic like above.
            store_dict1[int(dataframe.loc[i,"Octant"])]+=1
            if (int(dataframe.loc[i,"Octant"])!=1):
                store_dict1[1]=0
            if (int(dataframe.loc[i,"Octant"])!=2):
                store_dict1[2]=0
            if (int(dataframe.loc[i,"Octant"])!=3):
                store_dict1[3]=0
            if (int(dataframe.loc[i,"Octant"])!=4):
                store_dict1[4]=0
            if (int(dataframe.loc[i,"Octant"])!=-1):
                store_dict1[-1]=0
            if (int(dataframe.loc[i,"Octant"])!=-2):
                store_dict1[-2]=0
            if (int(dataframe.loc[i,"Octant"])!=-3):
                store_dict1[-3]=0
            if (int(dataframe.loc[i,"Octant"])!=-4):
                store_dict1[-4]=0
            #if present freq is equal to max freq then increase it by 1.
            if(store_dict1[int(dataframe.loc[i,"Octant"])]==max_store_dict[int(dataframe.loc[i,"Octant"])]):
                store_dict2[int(dataframe.loc[i,"Octant"])]+=1
                     #here when present_freq == max_freq then it means this is the end point of longest subseqnce, so i am storing the time at this row in the array.
                     #i am stroing it in the respective array accordingly.     
                if (int(dataframe.loc[i,"Octant"])==1):
                    range_list1.append(float(dataframe.loc[i,"T"]))
                elif (int(dataframe.loc[i,"Octant"])==2):
                    range_list2.append(float(dataframe.loc[i,"T"]))
                elif (int(dataframe.loc[i,"Octant"])==3):
                    range_list3.append(float(dataframe.loc[i,"T"]))
                elif (int(dataframe.loc[i,"Octant"])==4):
                    range_list4.append(float(dataframe.loc[i,"T"]))
                elif (int(dataframe.loc[i,"Octant"])==-1):
                    range_list_minus1.append(float(dataframe.loc[i,"T"]))
                elif (int(dataframe.loc[i,"Octant"])==-2):
                    range_list_minus2.append(float(dataframe.loc[i,"T"]))
                elif (int(dataframe.loc[i,"Octant"])==-3):
                    range_list_minus3.append(float(dataframe.loc[i,"T"]))
                elif (int(dataframe.loc[i,"Octant"])==-4):
                    range_list_minus4.append(float(dataframe.loc[i,"T"]))
        arr = ['+1','-1','+2','-2','+3','-3','+4','-4']
        # now printing the columns of count,longest subsequence length,count...from their dictonaries.
        for i in range(1,9):
            dataframe.loc[i,'             '] = max_store_dict[int(arr[i-1])]
            dataframe.loc[i,'              '] = store_dict2[int(arr[i-1])]

        dataframe['               '] = ''
        dataframe['Longest Subsquence Length with Range']=''
        dataframe['                '] = ''
        dataframe['                 '] = ''
        dataframe.loc[0,'Longest Subsquence Length with Range']='Octant###'
        dataframe.loc[0,'                ']='Longest Subsquence Length'
        dataframe.loc[0,'                 ']='Count'

        #this index maintains the current row in the output file.(where we have to print)
        present_row_var = 1
        for i in range(8):
            #here there are total 8divisions like +1,-1,+2,-2,+3,-3,+4,-4.
            flag = 0
            #with help of flag variable i have differentiated that in the first 2rows of ach division i should print from to etc
            #and in the remaining rows need to print from time to to time
            for xxx in range(store_dict2[int(arr[i])]+2):
                flag+=1
                if (flag==1):
                    dataframe.loc[present_row_var,'Longest Subsquence Length with Range'] = arr[i]
                    dataframe.loc[present_row_var,'                '] = max_store_dict[int(arr[i])]
                    dataframe.loc[present_row_var,'                 '] = store_dict2[int(arr[i])]
                elif (flag==2):
                    dataframe.loc[present_row_var,'Longest Subsquence Length with Range'] = "T"
                    dataframe.loc[present_row_var,'                '] = "From"
                    dataframe.loc[present_row_var,'                 '] = "To"
                else:
                    #column according to 1,-1,2,-2,3,-3,4,-4 i should take the respective array of time.
                    if (i==0):
                        #difference between each particular division is 0.01 so to find start time, we need to subtract the longest squbsqnce length*0.01 from end time.
                        dataframe.loc[present_row_var,'                '] = range_list1[flag-3] - (0.01)*(max_store_dict[int(arr[i])]-1)
                        #can store the directly array value as it is end time.
                        dataframe.loc[present_row_var,'                 '] = range_list1[flag-3]
                    elif (i==1):
                        dataframe.loc[present_row_var,'                '] = range_list_minus1[flag-3] - (0.01)*(max_store_dict[int(arr[i])]-1)
                        dataframe.loc[present_row_var,'                 '] = range_list_minus1[flag-3]                
                    elif (i==2):
                        dataframe.loc[present_row_var,'                '] = range_list2[flag-3] - (0.01)*(max_store_dict[int(arr[i])]-1)
                        dataframe.loc[present_row_var,'                 '] = range_list2[flag-3]
                    elif (i==3):
                        dataframe.loc[present_row_var,'                '] = range_list_minus2[flag-3] - (0.01)*(max_store_dict[int(arr[i])]-1)
                        dataframe.loc[present_row_var,'                 '] = range_list_minus2[flag-3]
                    elif (i==4):
                        dataframe.loc[present_row_var,'                '] = range_list3[flag-3] - (0.01)*(max_store_dict[int(arr[i])]-1)
                        dataframe.loc[present_row_var,'                 '] = range_list3[flag-3]
                    elif (i==5):
                        dataframe.loc[present_row_var,'                '] = range_list_minus3[flag-3] - (0.01)*(max_store_dict[int(arr[i])]-1)
                        dataframe.loc[present_row_var,'                 '] = range_list_minus3[flag-3]
                    elif (i==6):
                        dataframe.loc[present_row_var,'                '] = range_list4[flag-3] - (0.01)*(max_store_dict[int(arr[i])]-1)
                        dataframe.loc[present_row_var,'                 '] = range_list4[flag-3]
                    elif (i==7):
                        dataframe.loc[present_row_var,'                '] = range_list_minus4[flag-3] - (0.01)*(max_store_dict[int(arr[i])]-1)
                        dataframe.loc[present_row_var,'                 '] = range_list_minus4[flag-3]
                present_row_var+=1  
        # increase j by 1 as we are going to next row...   
        #             finally exporting this into output file by keeping index=False becuase we dont need the index column.
        #using the replace function, creating the output file according to input file.
        f2 = f.replace("input","output")
        f1= f2.replace('.xlsx'," cm_vel_octant_analysis_mod_"+str(mod)+".xlsx")
        dataframe.to_excel(f1,index=False)
        wb=openpyxl.load_workbook(f1)   
        ws=wb['Sheet1']

        #top,bottom,left,right is respective side border for the cell.
        top = Side(border_style='thin',color='000000')
        bottom = Side(border_style='thin',color='000000')
        left = Side(border_style='thin',color='000000')
        right = Side(border_style='thin',color='000000')
        border = Border(top=top,bottom=bottom,left=left,right=right)

        #ranges array stores the cell which i need to border.
        #according the format , i have made the ranges list .
        ranges=ws['N1':'N'+str(rows+2)]
        ranges1 = ws['O1':'O'+str(rows+2)]
        ranges2 = ws['P1':'P'+str(rows+2)]
        ranges3 = ws['Q1':'Q'+str(rows+2)]
        ranges4 = ws['R1':'R'+str(rows+2)]
        ranges5 = ws['S1':'S'+str(rows+2)]
        ranges6 = ws['T1':'T'+str(rows+2)]
        ranges7 = ws['U1':'U'+str(rows+2)]
        ranges8 = ws['V1':'V'+str(rows+2)]
        ranges9 = ws['W1':'W'+str(rows+2)]
        ranges10 = ws['X1':'X'+str(rows+2)]
        ranges11 = ws['Y1':'Y'+str(rows+2)]
        ranges12 = ws['Z1':'Z'+str(rows+2)]
        ranges13 = ws['AA1':'AA'+str(rows+2)]
        ranges14 = ws['AB1':'AB'+str(rows+2)]
        ranges15 = ws['AC1':'AC'+str(rows+2)]
        ranges16 = ws['AD1':'AD'+str(rows+2)]
        ranges17 = ws['AE1':'AE'+str(rows+2)]
        ranges18 = ws['AF1':'AF'+str(rows+2)]


        ranges19 = ws['AE'+str(rows+4):'AE'+str(rows+12)]
        ranges20 = ws['AC'+str(rows+4):'AC'+str(rows+12)]
        ranges21 = ws['AD'+str(rows+4):'AD'+str(rows+12)]

        ranges22 = ws['AS1':'AS10']
        ranges23 = ws['AU1':'AU10']
        ranges24 = ws['AT1':'AT10']

        ranges25 = ws['AW1':'AW'+str(present_row_var+1)]
        ranges26 = ws['AX1':'AX'+str(present_row_var+1)]
        ranges27 = ws['AY1':'AY'+str(present_row_var+1)]

        ranges28 = ws['AI1':'AI10']
        ranges29 = ws['AJ1':'AJ10']
        ranges30 = ws['AK1':'AK10']
        ranges31 = ws['AL1':'AL10']
        ranges32 = ws['AM1':'AM10']
        ranges33 = ws['AN1':'AN10']
        ranges34 = ws['AO1':'AO10']
        ranges35 = ws['AP1':'AP10']
        ranges36 = ws['AQ1':'AQ10']

        #now i am adding border to the each cell which present in the ranges array.
        for cell in ranges:
            for x in cell:
                x.border=border
        for cell in ranges1:
            for x in cell:
                x.border=border
        for cell in ranges2:
            for x in cell:
                x.border=border
        for cell in ranges3:
            for x in cell:
                x.border=border
        for cell in ranges4:
            for x in cell:
                x.border=border
        for cell in ranges5:
            for x in cell:
                x.border=border
        for cell in ranges6:
            for x in cell:
                x.border=border
        for cell in ranges7:
            for x in cell:
                x.border=border
        for cell in ranges8:
            for x in cell:
                x.border=border
        for cell in ranges9:
            for x in cell:
                x.border=border
        for cell in ranges10:
            for x in cell:
                x.border=border
        for cell in ranges11:
            for x in cell:
                x.border=border
        for cell in ranges12:
            for x in cell:
                x.border=border
        for cell in ranges13:
            for x in cell:
                x.border=border
        for cell in ranges14:
            for x in cell:
                x.border=border
        for cell in ranges15:
            for x in cell:
                x.border=border
        for cell in ranges16:
            for x in cell:
                x.border=border
        for cell in ranges17:
            for x in cell:
                x.border=border
        for cell in ranges18:
            for x in cell:
                x.border=border
        for cell in ranges19:
            for x in cell:
                x.border=border
        for cell in ranges20:
            for x in cell:
                x.border=border
        for cell in ranges21:
            for x in cell:
                x.border=border
        for cell in ranges22:
            for x in cell:
                x.border=border
        for cell in ranges23:
            for x in cell:
                x.border=border
        for cell in ranges24:
            for x in cell:
                x.border=border
        for cell in ranges25:
            for x in cell:
                x.border=border
        for cell in ranges26:
            for x in cell:
                x.border=border
        for cell in ranges27:
            for x in cell:
                x.border=border
        for cell in ranges28:
            for x in cell:
                x.border=border
        for cell in ranges29:
            for x in cell:
                x.border=border
        for cell in ranges30:
            for x in cell:
                x.border=border
        for cell in ranges31:
            for x in cell:
                x.border=border
        for cell in ranges32:
            for x in cell:
                x.border=border
        for cell in ranges33:
            for x in cell:
                x.border=border
        for cell in ranges34:
            for x in cell:
                x.border=border
        for cell in ranges35:
            for x in cell:
                x.border=border
        for cell in ranges36:
            for x in cell:
                x.border=border

        #here fill_cell object contains pattern/colour to fill in cell.
        fill_cell = PatternFill(patternType='solid',  fgColor='FFFF00') 

        for i in range(0,rows+1):
            #here i am storing the cell(rank octant value==1) as per the given in the question.
            if(dataframe.loc[i,'Rank Octant 1']==1):
                ws['W'+ str(i+2)].fill = fill_cell
            if(dataframe.loc[i,'Rank Octant 2']==1):
                ws['Y'+ str(i+2)].fill = fill_cell
            if(dataframe.loc[i,'Rank Octant 3']==1):
                ws['AA'+ str(i+2)].fill = fill_cell
            if(dataframe.loc[i,'Rank Octant 4']==1):
                ws['AC'+ str(i+2)].fill = fill_cell
            if(dataframe.loc[i,'Rank Octant -1']==1):
                ws['X'+ str(i+2)].fill = fill_cell
            if(dataframe.loc[i,'Rank Octant -2']==1):
                ws['Z'+ str(i+2)].fill = fill_cell
            if(dataframe.loc[i,'Rank Octant -3']==1):
                ws['AB'+ str(i+2)].fill = fill_cell
            if(dataframe.loc[i,'Rank Octant -4']==1):
                ws['AD'+ str(i+2)].fill = fill_cell

        for i in range(0,len(colour_row)):
            #here in the tranisiton rows(overall and also mod transition part..)
            maxi_maxi = 0
            #above variable stores max value in the row(columns or AJ....AQ (when transition valye present ))
            maxi_maxi = max(maxi_maxi,ws['AJ'+str(colour_row[i])].value)
            maxi_maxi = max(maxi_maxi,ws['AK'+str(colour_row[i])].value)
            maxi_maxi = max(maxi_maxi,ws['AL'+str(colour_row[i])].value)
            maxi_maxi = max(maxi_maxi,ws['AM'+str(colour_row[i])].value)
            maxi_maxi = max(maxi_maxi,ws['AN'+str(colour_row[i])].value)
            maxi_maxi = max(maxi_maxi,ws['AO'+str(colour_row[i])].value)
            maxi_maxi = max(maxi_maxi,ws['AP'+str(colour_row[i])].value)
            maxi_maxi = max(maxi_maxi,ws['AQ'+str(colour_row[i])].value)
            #now using if conditions, colouring the cell if it is equal to  maxi_maxi
            if (maxi_maxi==ws['AJ'+str(colour_row[i])].value):
                ws['AJ'+str(colour_row[i])].fill = fill_cell
            if (maxi_maxi==ws['AK'+str(colour_row[i])].value):
                ws['AK'+str(colour_row[i])].fill = fill_cell
            if (maxi_maxi==ws['AL'+str(colour_row[i])].value):
                ws['AL'+str(colour_row[i])].fill = fill_cell     
            if (maxi_maxi==ws['AM'+str(colour_row[i])].value):
                ws['AM'+str(colour_row[i])].fill = fill_cell    
            if (maxi_maxi==ws['AN'+str(colour_row[i])].value):
                ws['AN'+str(colour_row[i])].fill = fill_cell
            if (maxi_maxi==ws['AO'+str(colour_row[i])].value):
                ws['AO'+str(colour_row[i])].fill = fill_cell
            if (maxi_maxi==ws['AP'+str(colour_row[i])].value):
                ws['AP'+str(colour_row[i])].fill = fill_cell
            if (maxi_maxi==ws['AQ'+str(colour_row[i])].value):
                ws['AQ'+str(colour_row[i])].fill = fill_cell
        #here two in overall transition box appling border to the each cell.
        for i in range(17,len(dataframe),14):
            if(dataframe.loc[i-2,'Overall Transition Count']=='+1'):
                ranges37 = ws['AI'+str(i):'AI'+str(i+7)]
                for cell in ranges37:
                    for x in cell:
                        x.border=border
                ranges38 = ws['AJ'+str(i):'AJ'+str(i+7)]
                for cell in ranges38:
                    for x in cell:
                        x.border=border
                ranges39 = ws['AK'+str(i):'AK'+str(i+7)]
                for cell in ranges39:
                    for x in cell:
                        x.border=border
                ranges40 = ws['AL'+str(i):'AL'+str(i+7)]
                for cell in ranges40:
                    for x in cell:
                        x.border=border
                ranges41 = ws['AM'+str(i):'AM'+str(i+7)]
                for cell in ranges41:
                    for x in cell:
                        x.border=border
                ranges42 = ws['AN'+str(i):'AN'+str(i+7)]
                for cell in ranges42:
                    for x in cell:
                        x.border=border
                ranges43 = ws['AO'+str(i):'AO'+str(i+7)]
                for cell in ranges43:
                    for x in cell:
                        x.border=border
                ranges44 = ws['AP'+str(i):'AP'+str(i+7)]
                for cell in ranges44:
                    for x in cell:
                        x.border=border
                ranges45 = ws['AQ'+str(i):'AQ'+str(i+7)]
                for cell in ranges45:
                    for x in cell:
                        x.border=border
            else:
                break

        wb.save(f1)
        print("Your output file for the above input file is ready and named as "+ f1)
        print("\n")
    ##Read all the excel files in a batch format from the input/ folder. Only xlsx to be allowed
    ##Save all the excel files in a the output/ folder. Only xlsx to be allowed
    ## output filename = input_filename[_octant_analysis_mod_5000].xlsx , ie, append _octant_analysis_mod_5000 to the original filename. 

    ###Code
from platform import python_version
ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

#5000
mod=5000
octant_analysis(mod)
#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))


